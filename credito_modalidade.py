# %% [markdown]
# # PRIMEIRA PARTE - downloads manual dos arquivos do INVESTSUS

# %%
# 📚 BIBLIOTECAS
import os
import time
import warnings
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.edge.service import Service
from selenium.webdriver.edge.options import Options

# 🔕 Oculta alertas
warnings.filterwarnings('ignore')

# Contagem de TEMPO de processamento
inicio = datetime.now()
print(f"🔵 Início da execução: {inicio.strftime('%H:%M:%S')}")

# 📁 Diretório de downloads
DOWNLOAD_DIR = os.path.join(os.getcwd(), "downloads")
os.makedirs(DOWNLOAD_DIR, exist_ok=True)
print(f"📁 Diretório de downloads configurado: {DOWNLOAD_DIR}")

# ⚙️ Configuração do Edge
driver_path = os.path.join(os.getcwd(), "web", "msedgedriver.exe")
service = Service(executable_path=driver_path)
edge_options = Options()

# Configura o modo headless (invisível)
edge_options.add_argument("--headless")             # Executa sem abrir a janela
edge_options.add_argument("--disable-gpu")          # Evita problemas gráficos
edge_options.add_argument("--window-size=1920,1080")  # Define tamanho da janela virtual

# Configurações de download
edge_options.add_experimental_option("prefs", {
    "download.default_directory": DOWNLOAD_DIR,
    "download.prompt_for_download": False,
    "directory_upgrade": True,
    "safebrowsing.enabled": True
})

# 🚀 Inicializa o navegador com as opções já configuradas
driver = webdriver.Edge(service=service, options=edge_options)

wait = WebDriverWait(driver, 20)
print("🚀 Navegador Edge iniciado com sucesso.")

# Configura o modo headless
edge_options.add_argument("--headless")  # Executa sem abrir a janela
edge_options.add_argument("--disable-gpu")  # Evita problemas gráficos
edge_options.add_argument("--window-size=1920,1080")  # Define tamanho da janela virtual


# 🌐 Acessa a página alvo
url = "https://investsuspaineis.saude.gov.br/extensions/CGIN_PMAE/CGIN_PMAE.html#"
driver.get(url)
print(f"🌐 Página acessada: {url}")
time.sleep(5)

# 🔁 Função para baixar e renomear arquivos
def baixar_e_renomear(xpath_botao, nome_destino):
    print(f"📥 Iniciando download para: {nome_destino}")
    
    # Clica no botão de download
    botao = wait.until(EC.element_to_be_clickable((By.XPATH, xpath_botao)))
    botao.click()
    time.sleep(2)

    # Aceita o alerta
    WebDriverWait(driver, 10).until(EC.alert_is_present())
    alerta = driver.switch_to.alert
    alerta.accept()

    # Aguarda o download finalizar
    time.sleep(5)

    # Renomeia o arquivo mais recente .xlsx
    arquivos_xlsx = [os.path.join(DOWNLOAD_DIR, f) for f in os.listdir(DOWNLOAD_DIR) if f.endswith(".xlsx")]
    arquivo_mais_recente = max(arquivos_xlsx, key=os.path.getctime)
    caminho_novo = os.path.join(DOWNLOAD_DIR, nome_destino)

    # Se já existir um arquivo com o nome de destino, exclui
    if os.path.exists(caminho_novo):
        os.remove(caminho_novo)

    # Renomeia o novo arquivo
    os.rename(arquivo_mais_recente, caminho_novo)
    print(f"📦 Arquivo renomeado para: {nome_destino}\n")


# 📊 Aba Crédito
print("📂 Acessando aba: Crédito Financeiro")
aba_credito = wait.until(EC.element_to_be_clickable((By.XPATH,'//*[@id="menu_abas"]/a[3]')))
aba_credito.click()
time.sleep(5)

baixar_e_renomear('//*[@id="QV3-02574e8688-0e17-49d8-8ca9-c037abb4a5f7"]', "credito_financeiro_aba1.xlsx")
baixar_e_renomear('//*[@id="QV3-03a27c0e0b-cac7-45c5-92c1-ec1cb34f3828"]', "credito_financeiro_aba2.xlsx")
baixar_e_renomear('//*[@id="QV3-04541ab7f8-9dda-4cbd-82cf-72840cb4ac2d"]', "credito_financeiro_aba3.xlsx")

# 📊 Aba Modalidade 1
print("📂 Acessando aba: Modalidade 1")
aba_modalidade = wait.until(EC.element_to_be_clickable((By.XPATH,'//*[@id="menu_abas"]/a[4]')))
aba_modalidade.click()
time.sleep(5)

baixar_e_renomear('//*[@id="QV4-02yfhvCp"]', "modalidade_1_aba1.xlsx")
baixar_e_renomear('//*[@id="QV4-03JQRjW"]', "modalidade_1_aba2.xlsx")
baixar_e_renomear('//*[@id="QV4-04gvqxmPC"]', "modalidade_1_aba3.xlsx")

driver.quit()


# %% [markdown]
# # SEGUNDA PARTE, REALIZAR TRATAMENTO DOS DADOS.

# %%

import pandas as pd
import numpy as np
# Ajusta a opção de exibição para mostrar todas as colunas
pd.set_option('display.max_columns', None)


# Créditos financeiros
df_cf_aba1 = pd.read_excel("downloads/credito_financeiro_aba1.xlsx")
df_cf_aba2 = pd.read_excel("downloads/credito_financeiro_aba2.xlsx")
df_cf_aba3 = pd.read_excel("downloads/credito_financeiro_aba3.xlsx")

# Modalidade 1
df_m1_aba1 = pd.read_excel("downloads/modalidade_1_aba1.xlsx")
df_m1_aba2 = pd.read_excel("downloads/modalidade_1_aba2.xlsx")
df_m1_aba3 = pd.read_excel("downloads/modalidade_1_aba3.xlsx")


# %%
# Lista de propostas a aprovar
propostas_aprovada = [
    781831300012025501   # APROVAÇÃO DE SOBRAL (MANUALMENTE)
]

# Converte os valores da lista para string
propostas_aprovada = [str(p) for p in propostas_aprovada]

# Atualiza o status nos três DataFrames
for i, df in enumerate([df_cf_aba1, df_cf_aba2, df_cf_aba3], start=1):
    # Garante que a coluna esteja no tipo string
    df['Proposta de Referência'] = df['Proposta de Referência'].astype(str)
    
    # Aplica a alteração de status
    df.loc[df['Proposta de Referência'].isin(propostas_aprovada), 'Status da Proposta'] = 'Aprovado'
    
    # Verifica quantas foram alteradas
    aprovadas = df[df['Proposta de Referência'].isin(propostas_aprovada)]
    print(f"✅ df_cf_aba{i}: {len(aprovadas)} propostas marcadas como 'Aprovado'.")


# %%
# Lista de propostas a cancelar
propostas_canceladas = [
    1086978200012025503,    # PE-RECIFE 
    8862568600242025502,    # RS-PORTO ALEGRE
    2870053000032025501,    # SC-TIMBE DO SUL
    2870053000022025502,    # SC-SOMBRIO
    2870053000022025503,    # SC-SOMBRIO
    4605648700012025501,    # SP-VALINHOS
    24712500022025504,      # RJ-NITEROI
    353343200012025502,     # CAMPO GRANDE
]

# Converte os valores da lista para string
propostas_canceladas = [str(p) for p in propostas_canceladas]

# Atualiza o status nos três DataFrames
for i, df in enumerate([df_cf_aba1, df_cf_aba2, df_cf_aba3], start=1):
    # Garante que a coluna esteja no tipo string
    df['Proposta de Referência'] = df['Proposta de Referência'].astype(str)
    
    # Aplica a alteração de status
    df.loc[df['Proposta de Referência'].isin(propostas_canceladas), 'Status da Proposta'] = 'Cancelado'
    
    # Verifica quantas foram alteradas
    canceladas = df[df['Proposta de Referência'].isin(propostas_canceladas)]
    print(f"✅ df_cf_aba{i}: {len(canceladas)} propostas marcadas como 'Cancelado'.")


# %%
# Mapeamento CNPJ → CNES
CNPJ_CNES = {
    5048983000150: 3151700,
    85514370000108: 3021238,
    80906639000170: 4055748,
    5089379000171: 2415739,
    1273401000188: 3025020,
    45184066000117: 3042529,
    72551799000115: 2080281,
    9407153000122: 6012302
}

# DataFrames com CNPJ
df_com_cnpj = [df_cf_aba1, df_m1_aba1]

# Atualizar CNES apenas onde CNPJ está no dicionário
for df in df_com_cnpj:
    df['CNPJ'] = df['CNPJ'].astype(int)
    df['CNES'] = np.where(
        df['CNPJ'].isin(CNPJ_CNES),
        df['CNPJ'].map(CNPJ_CNES),
        df['CNES']
    ).astype('object')

# Criar dicionário Proposta → CNES a partir dos DataFrames atualizados
proposta_cnes_map = pd.concat(df_com_cnpj)[['Proposta de Referência', 'CNES']].dropna()
proposta_cnes_dict = dict(zip(proposta_cnes_map['Proposta de Referência'], proposta_cnes_map['CNES']))

# DataFrames sem CNPJ
df_sem_cnpj = [df_cf_aba2, df_cf_aba3, df_m1_aba2, df_m1_aba3]

# Atualizar CNES apenas onde Proposta de Referência está no dicionário
for df in df_sem_cnpj:
    df['CNES'] = np.where(
        df['Proposta de Referência'].isin(proposta_cnes_dict),
        df['Proposta de Referência'].map(proposta_cnes_dict),
        df['CNES']
    ).astype('object')


# %%
# conversão modalidade 1

for i, df in enumerate([df_m1_aba1, df_m1_aba2, df_m1_aba3], start=1):
    # Garante que a coluna esteja no tipo string
    df['Proposta de Referência'] = df['Proposta de Referência'].astype(str)


# %%
# montando a MATRIZ DE OFERTA - CRÉDITO FINANCEIRO - CIRURGIAS

# remover coluna
df_cf_aba3.drop(columns='TP_COMPLEXIDADE', inplace=True)


# colocar coluna 'entidade'
df_cf_aba3 = df_cf_aba3.merge(df_cf_aba1[['Proposta de Referência', 'Entidade']],                                   
    on='Proposta de Referência',
    how='left')

# alterando nome da colunas
df_cf_aba3.rename(columns={
    'Entidade': 'ENTIDADE','TX_COMPLEMENTACAO_MAXIMA': '% COMPLEMENTACAO_MAXIMA'                                        
}, inplace=True)

# reoganizado as colunas
df_cf_aba3 = df_cf_aba3[[                                                                                               
    'Proposta de Referência',
    'Status da Proposta',
    'UF',
    'Município',
    'CNES',
    'ENTIDADE',
    'CO_PROCEDIMENTO_SIGTAP',
    'NO_GRUPO',
    'NO_PROCEDIMENTO',
    '% COMPLEMENTACAO_MAXIMA',
    'VL_TABELA_SUS',
    'VL_TOTAL_COMPLEMENTACAO_MAXIMA',
    'VL_MEDIA_BRASIL_CALCULADO',
    'QT_ATENDIMENTO_MES',
    'VL_TOTAL'
]]


# %%
df_cf_aba3.info()

# %%
# montando a MATRIZ DE OFERTA - MODALIDADE 1 - CIRURGIAS

# remover coluna 'TP_COMPLEXIDADE'
df_m1_aba3.drop(columns='TP_COMPLEXIDADE', inplace=True)

# remover coluna 'Entidade' existente, se houver
if 'Entidade' in df_m1_aba3.columns:
    df_m1_aba3.drop(columns='Entidade', inplace=True)

# colocar coluna 'Entidade' via merge
df_m1_aba3 = df_m1_aba3.merge(
    df_m1_aba1[['Proposta de Referência', 'Entidade']],
    on='Proposta de Referência',
    how='left'
)

# renomear colunas
df_m1_aba3.rename(columns={
    'Entidade': 'ENTIDADE',
    'TX_COMPLEMENTACAO_MAXIMA': '% COMPLEMENTACAO_MAXIMA'
}, inplace=True)

# reorganizar colunas
df_m1_aba3 = df_m1_aba3[[                                                                                               
    'Proposta de Referência',
    'Status da Proposta',
    'UF',
    'Município',
    'CNES',
    'ENTIDADE',
    'CO_PROCEDIMENTO_SIGTAP',
    'NO_GRUPO',
    'NO_PROCEDIMENTO',
    '% COMPLEMENTACAO_MAXIMA',
    'VL_TABELA_SUS',
    'VL_TOTAL_COMPLEMENTACAO_MAXIMA',
    'VL_MEDIA_BRASIL_CALCULADO',
    'QT_ATENDIMENTO_MES',
    'VL_TOTAL'
]]


# %%
df_m1_aba3.info()

# %%
# montando a MATRIZ DE OFERTA - CRÉDITO FINANCEIRO - OCI

# remover coluna
df_cf_aba2.drop(columns=['TP_SEXO','NU_IDADE_MINIMA','NU_IDADE_MAXIMA','VL_MEDIA_BRASIL_CALCULADO'], inplace=True)                   


# colocar coluna 'entidade'
df_cf_aba2 = df_cf_aba2.merge(df_cf_aba1[['Proposta de Referência', 'Entidade']],                        
    on='Proposta de Referência',
    how='left')


# reoganizado as colunas
df_cf_aba2= df_cf_aba2[[                                                                                 
    'Proposta de Referência',
    'Status da Proposta',
    'UF',
    'Município',
    'CNES',
    'Entidade',
    'NU_PROCEDIMENTO',
    'NO_GRUPO',
    'NO_PROCEIDMENTO',
    'DS_PROCEDIMENTO',
    'QT_ATENDIMENTO_MES',
    'VL_CALCULADO',
    'VL_TOTAL'
]]


# %%
# montando a MATRIZ DE OFERTA - MODALIDADE 1 - OCI



# Remover colunas desnecessárias
df_m1_aba2.drop(columns=['TP_SEXO', 'NU_IDADE_MINIMA', 'NU_IDADE_MAXIMA'], inplace=True)

# Criar coluna de valor total
df_m1_aba2['VALOR_TOTAL_MES'] = df_m1_aba2['QT_ATENDIMENTO_MES'] * df_m1_aba2['VL_PROCEDIMENTO']

# Adicionar coluna 'Entidade' via merge
df_m1_aba2 = df_m1_aba2.merge(
    df_m1_aba1[['Proposta de Referência', 'Entidade']],
    on='Proposta de Referência',
    how='left'
)

# Reorganizar colunas
df_m1_aba2 = df_m1_aba2[[
    'Proposta de Referência',
    'Status da Proposta',
    'UF',
    'Município',
    'CNES',
    'Entidade',
    'NU_PROCEDIMENTO',
    'NO_GRUPO',
    'NO_PROCEIDMENTO',
    'DS_PROCEDIMENTO',
    'QT_ATENDIMENTO_MES',
    'VL_CALCULADO',
    'VL_TOTAL'
]]



# %%
# montando a aba SIMPLIFICADA - CRÉDITO FINANCEIRO 


# copiando a informação da df_cf_aba1
df_simp_cc = df_cf_aba1.copy()

# Remover colunas desnecessárias
df_simp_cc.drop(columns=[
    'Dt. Cadastro', 'Dt. Atualização', 'Dívida Aprox.', 'VL_SALDO_DEVEDOR', 
    'VL_TRIBUTO_FEDERAL_ESTIMADO'], inplace=True)

# Converter a coluna VL_TOTAL para numérico (float), tratando erros
df_cf_aba3['VL_TOTAL'] = pd.to_numeric(df_cf_aba3['VL_TOTAL'], errors='coerce')

# Agrupar df_cf_aba3 por 'Proposta de Referência' e somar 'VALOR_TOTAL_MES' --- CIRURGIAS
soma_por_proposta_cc = df_cf_aba3.groupby('Proposta de Referência')['VL_TOTAL'].sum().reset_index()

# Renomear a coluna para o nome desejado
soma_por_proposta_cc.rename(columns={'VL_TOTAL': 'VL_TOTAL_COMP_CIRUGICO'}, inplace=True)

# Fazer o merge com df_simp_cc
df_simp_cc = df_simp_cc.merge(soma_por_proposta_cc, on='Proposta de Referência', how='left')


# Agrupar df_cf_aba2 por 'Proposta de Referência' e somar 'VALOR_TOTAL_MES' --- OCI
soma_por_proposta_co = df_cf_aba2.groupby('Proposta de Referência')['VL_TOTAL'].sum().reset_index()

# Renomear a coluna para o nome desejado
soma_por_proposta_co.rename(columns={'VL_TOTAL': 'VL_TOTAL_OCI'}, inplace=True)

# Fazer o merge com df_simp_cc
df_simp_cc = df_simp_cc.merge(soma_por_proposta_co, on='Proposta de Referência', how='left')

# %%
# nova coluna de total COMP + OCI
df_simp_cc['VL_TOTAL_COMP_CIRUGICO'].fillna(0, inplace=True)
df_simp_cc['VL_TOTAL_OCI'].fillna(0, inplace=True)
df_simp_cc['VALOR_TOTAL_MES_COMP+OCI'] = df_simp_cc['VL_TOTAL_COMP_CIRUGICO'] + df_simp_cc['VL_TOTAL_OCI']
df_simp_cc['VALOR_TOTAL_ANO_COMP+OCI']= df_simp_cc['VALOR_TOTAL_MES_COMP+OCI']*12


# %%
# aba de Cancelado
df_proposta_cancelada = df_simp_cc[df_simp_cc['Status da Proposta'] == 'Cancelado'].copy()


# %%
# Montando a aba SIMPLIFICADA - MODALIDADE 1

# Copiar os dados da aba 1
df_simp_m1 = df_m1_aba1.copy()

# Remover colunas desnecessárias
df_simp_m1.drop(columns=['Dt. Cadastro', 'Dt. Atualização'], inplace=True)

# Converter a coluna VL_TOTAL de df_m1_aba3 para numérico
df_m1_aba3['VL_TOTAL'] = pd.to_numeric(df_m1_aba3['VL_TOTAL'], errors='coerce')

# Agrupar df_m1_aba3 por 'Proposta de Referência' e somar VL_TOTAL (CIRURGIAS)
soma_por_proposta_mc = df_m1_aba3.groupby('Proposta de Referência')['VL_TOTAL'].sum().reset_index()
soma_por_proposta_mc.rename(columns={'VL_TOTAL': 'VL_TOTAL_COMP_CIRUGICO'}, inplace=True)

# Merge com df_simp_m1
df_simp_m1 = df_simp_m1.merge(soma_por_proposta_mc, on='Proposta de Referência', how='left')

# Converter a coluna VL_TOTAL de df_m1_aba2 para numérico (caso necessário)
df_m1_aba2['VL_TOTAL'] = pd.to_numeric(df_m1_aba2['VL_TOTAL'], errors='coerce')

# Agrupar df_m1_aba2 por 'Proposta de Referência' e somar VL_TOTAL (OCI)
soma_por_proposta_mo = df_m1_aba2.groupby('Proposta de Referência')['VL_TOTAL'].sum().reset_index()
soma_por_proposta_mo.rename(columns={'VL_TOTAL': 'VL_TOTAL_OCI'}, inplace=True)

# Merge final com df_simp_m1
df_simp_m1 = df_simp_m1.merge(soma_por_proposta_mo, on='Proposta de Referência', how='left')


# %%
# nova coluna de total COMP + OCI
df_simp_m1['VL_TOTAL_COMP_CIRUGICO'].fillna(0, inplace=True)
df_simp_m1['VL_TOTAL_OCI'].fillna(0, inplace=True)
df_simp_m1['VALOR_TOTAL_MES_COMP+OCI'] = df_simp_m1['VL_TOTAL_COMP_CIRUGICO'] + df_simp_m1['VL_TOTAL_OCI']
df_simp_m1['VALOR_TOTAL_ANO_COMP+OCI']= df_simp_m1['VALOR_TOTAL_MES_COMP+OCI']*12

# %% [markdown]
# # TERCEIRA PARTE, Carregar os dataFrame para a tabela MODELO

# %%
import os
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

# Diretórios e arquivos
MODELO_FILENAME = "MONITORAMENTO DE COMPONENTE.xlsx"
MODELO_DIR = os.path.join(os.getcwd(), "model")
MODELO_PATH = os.path.join(MODELO_DIR, MODELO_FILENAME)

# Mapeamento dos arquivos para as abas
MAPPING = {
    "df_cf_aba1": "CREDITO_FINANCEIRO",
    "df_cf_aba2": "M_OFERTA_CF_OCI",
    "df_cf_aba3": "M_OFERTA_CF_CC",
    "df_simp_cc": "SIMP_CF",
    "df_m1_aba1": "MODALIDADE_1",
    "df_m1_aba2": "M_OFERTA_M1_OCI",
    "df_m1_aba3": "M_OFERTA_M1_CC",
    "df_simp_m1": "SIMP_M1",
    "df_proposta_cancelada": "CCPP-CANCELAR",
}

# Função para sobrescrever a aba a partir da linha 3
def sobrescrever_aba(workbook, aba_nome, df):
    if aba_nome in workbook.sheetnames:
        ws = workbook[aba_nome]
        for i, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=3):
            for j, value in enumerate(row, start=1):
                ws.cell(row=i, column=j, value=value)
        print(f"✅ Aba '{aba_nome}' atualizada com {len(df)} linhas.")
    else:
        print(f"⚠️ Aba '{aba_nome}' não encontrada no modelo.")

# Função para carregar os DataFrames (simplesmente acessa variáveis globais)
def carregar_dados_do_excel(nome_df):
    try:
        return globals()[nome_df]
    except KeyError:
        print(f"⚠️ DataFrame '{nome_df}' não está definido.")
        return None

# Execução principal
if not os.path.exists(MODELO_PATH):
    print(f"❌ Erro: O arquivo modelo esperado '{MODELO_FILENAME}' não foi encontrado em '{MODELO_DIR}'.")
    print("Verifique se o caminho do arquivo modelo está correto.")
else:
    try:
        modelo_wb = openpyxl.load_workbook(MODELO_PATH)
        print("📂 Arquivo modelo carregado.")

        # Atualizar abas conforme mapeamento
        for df_nome, aba_nome in MAPPING.items():
            df = carregar_dados_do_excel(df_nome)
            if df is not None:
                sobrescrever_aba(modelo_wb, aba_nome, df)

        # Atualizar aba INFO
        print("\n🛠 Atualizando aba 'INFO'...")
        sigla_uf = "BR"

        if 'INFO' in modelo_wb.sheetnames:
            aba_info = modelo_wb['INFO']
            aba_info['H2'] = datetime.now().strftime('%d/%m/%Y %H:%M')
            aba_info['G1'] = sigla_uf
            print(f"   -> Aba 'INFO' atualizada (H2: Data/Hora, G1: {sigla_uf}).")
        else:
            print("   ⚠️ Aba 'INFO' não encontrada no arquivo modelo.")

        # Salvar novo arquivo
        os.makedirs("saida", exist_ok=True)
        novo_nome = os.path.join("saida", f"{datetime.today().strftime('%Y%m%d')}_{MODELO_FILENAME}")
        modelo_wb.save(novo_nome)

        print("\n-----------------------------------------------------------------")
        print(f"🎉 Sucesso! O novo arquivo '{novo_nome}' foi criado.")
        print("   As abas 'VISÃO_GERAL' e 'VISÃO_GERAL_M1' devem ter sido recalculadas pelo Excel.")
        print("-----------------------------------------------------------------")

    except Exception as e:
        print(f"❌ Erro fatal durante a execução: {e}")


# %%
# Caminho da pasta Downloads
pasta = os.path.expanduser("~/Downloads")

# Percorre todos os arquivos na pasta
for arquivo in os.listdir(pasta):
    caminho_arquivo = os.path.join(pasta, arquivo)
    if os.path.isfile(caminho_arquivo):
        os.remove(caminho_arquivo)


# %%
fim = datetime.now()
tempo_total = fim - inicio

horas, resto = divmod(tempo_total.total_seconds(), 3600)
minutos, segundos = divmod(resto, 60)

print(f"✅ Tempo total de execução: {int(horas)}h {int(minutos)}min {int(segundos)}s")


